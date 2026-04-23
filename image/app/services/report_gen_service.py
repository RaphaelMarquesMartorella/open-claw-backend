'''Geracao de relatorios em PDF e Excel.'''
import base64
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.ext.asyncio import AsyncSession

from repositories import (
    get_clientes_inativos_repository,
    get_comparativo_diario_repository,
    get_kpis_repository,
    get_top_produtos_repository,
    get_vendedor_ranking_repository,
)

PDF_MIME = 'application/pdf'
XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_ALIGN = Alignment(horizontal='center')
HEADER_COLOR = colors.HexColor('#1F4E79')
STRIPE_COLOR = colors.HexColor('#F2F2F2')


def _fmt_brl(v: float | int) -> str:
    return f'R$ {v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def _fmt_pct(v: float | int) -> str:
    return f'{v:+.1f}%'


def _today_tag() -> str:
    return datetime.now(ZoneInfo('America/Recife')).strftime('%Y%m%d_%H%M')


def _table_style() -> TableStyle:
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_COLOR),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, STRIPE_COLOR]),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])


def _apply_excel_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _auto_width(ws) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


async def _pdf_doc(title: str) -> tuple[BytesIO, SimpleDocTemplate, list]:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story: list = [
        Paragraph(f'<b>{title}</b>', styles['Title']),
        Paragraph(datetime.now(ZoneInfo('America/Recife')).strftime('Gerado em %d/%m/%Y %H:%M'), styles['Italic']),
        Spacer(1, 12),
    ]
    return buf, doc, story


async def _ranking_pdf(db: AsyncSession) -> bytes:
    ranking = await get_vendedor_ranking_repository(db)
    buf, doc, story = await _pdf_doc('Ranking de Vendedores')
    data = [['Cod', 'Apelido', 'Fat. Atual', 'Fat. Anterior', 'Delta %', 'Pedidos']]
    for v in ranking:
        data.append([
            v['codvend'], v['apelido'],
            _fmt_brl(v['fat_atual']), _fmt_brl(v['fat_anterior']),
            _fmt_pct(v['delta_percentual']), v.get('pedidos_atual', 0),
        ])
    story.append(Table(data, repeatRows=1, style=_table_style()))
    doc.build(story)
    return buf.getvalue()


async def _ranking_xlsx(db: AsyncSession) -> bytes:
    ranking = await get_vendedor_ranking_repository(db)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ranking'
    ws.append(['Cod', 'Apelido', 'Fat. Atual', 'Fat. Anterior', 'Delta %', 'Pedidos'])
    for v in ranking:
        ws.append([v['codvend'], v['apelido'], v['fat_atual'], v['fat_anterior'], v['delta_percentual'] / 100, v.get('pedidos_atual', 0)])
    _apply_excel_header(ws, 6)
    for cell in ws['C'][1:]:
        cell.number_format = 'R$ #,##0.00'
    for cell in ws['D'][1:]:
        cell.number_format = 'R$ #,##0.00'
    for cell in ws['E'][1:]:
        cell.number_format = '0.0%'
    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _produtos_pdf(db: AsyncSession) -> bytes:
    produtos = await get_top_produtos_repository(db, limit=30)
    buf, doc, story = await _pdf_doc('Top Produtos por Faturamento')
    data = [['Cod', 'Descricao', 'Marca', 'Qtd', 'Faturamento', 'Estoque']]
    for p in produtos:
        data.append([
            p['codprod'],
            (p['descrprod'] or '')[:40],
            p.get('marca', '') or '',
            int(p['qtd_vendida']),
            _fmt_brl(p['faturamento']),
            int(p.get('estoque', 0) or 0),
        ])
    story.append(Table(data, repeatRows=1, style=_table_style()))
    doc.build(story)
    return buf.getvalue()


async def _produtos_xlsx(db: AsyncSession) -> bytes:
    produtos = await get_top_produtos_repository(db, limit=50)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Top Produtos'
    ws.append(['Cod', 'Descricao', 'Marca', 'Qtd Vendida', 'Faturamento', 'Estoque', 'Giro'])
    for p in produtos:
        ws.append([
            p['codprod'], p['descrprod'], p.get('marca'),
            float(p['qtd_vendida']), float(p['faturamento']),
            int(p.get('estoque') or 0), float(p.get('giro') or 0),
        ])
    _apply_excel_header(ws, 7)
    for cell in ws['E'][1:]:
        cell.number_format = 'R$ #,##0.00'
    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _inativos_pdf(db: AsyncSession) -> bytes:
    clientes = await get_clientes_inativos_repository(db)
    buf, doc, story = await _pdf_doc('Clientes Inativos no Mes')
    data = [['Cod', 'Cliente', 'Cidade/UF', 'Valor Anterior', 'Vendedor']]
    for c in clientes:
        data.append([
            c['codparc'],
            (c['nomeparc'] or '')[:30],
            f"{c.get('cidade', '') or ''}/{c.get('uf', '') or ''}",
            _fmt_brl(c['valor_anterior']),
            c.get('apelido') or '',
        ])
    story.append(Table(data, repeatRows=1, style=_table_style()))
    doc.build(story)
    return buf.getvalue()


async def _inativos_xlsx(db: AsyncSession) -> bytes:
    clientes = await get_clientes_inativos_repository(db)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inativos'
    ws.append(['Cod', 'Cliente', 'Cidade', 'UF', 'Valor Anterior', 'Vendedor'])
    for c in clientes:
        ws.append([
            c['codparc'], c['nomeparc'], c.get('cidade'), c.get('uf'),
            float(c['valor_anterior']), c.get('apelido'),
        ])
    _apply_excel_header(ws, 6)
    for cell in ws['E'][1:]:
        cell.number_format = 'R$ #,##0.00'
    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _diario_pdf(db: AsyncSession, dias: int = 30) -> bytes:
    rows = await get_comparativo_diario_repository(db, dias)
    buf, doc, story = await _pdf_doc(f'Faturamento Diario (ultimos {dias} dias)')
    data = [['Data', 'Faturamento']]
    for r in rows:
        data.append([r['data'], _fmt_brl(r['faturamento'])])
    story.append(Table(data, repeatRows=1, style=_table_style()))
    doc.build(story)
    return buf.getvalue()


async def _diario_xlsx(db: AsyncSession, dias: int = 30) -> bytes:
    rows = await get_comparativo_diario_repository(db, dias)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Diario'
    ws.append(['Data', 'Faturamento'])
    for r in rows:
        ws.append([r['data'], float(r['faturamento'])])
    _apply_excel_header(ws, 2)
    for cell in ws['B'][1:]:
        cell.number_format = 'R$ #,##0.00'
    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _completo_xlsx(db: AsyncSession) -> bytes:
    kpis = await get_kpis_repository(db)
    ranking = await get_vendedor_ranking_repository(db)
    produtos = await get_top_produtos_repository(db, limit=50)
    inativos = await get_clientes_inativos_repository(db)
    diario = await get_comparativo_diario_repository(db, 30)

    wb = Workbook()

    ws = wb.active
    ws.title = 'KPIs'
    ws.append(['Metrica', 'Valor'])
    ws.append(['Faturamento Atual', kpis['faturamento_atual']])
    ws.append(['Faturamento Anterior', kpis['faturamento_anterior']])
    ws.append(['Delta %', kpis['delta_percentual'] / 100])
    ws.append(['Pedidos Atual', kpis.get('pedidos_atual', 0)])
    ws.append(['Pedidos Anterior', kpis.get('pedidos_anterior', 0)])
    ws.append(['Clientes Ativos', kpis.get('clientes_ativos', 0)])
    _apply_excel_header(ws, 2)
    for cell in ws['B'][1:3]:
        cell.number_format = 'R$ #,##0.00'
    ws['B4'].number_format = '0.0%'
    _auto_width(ws)

    ws2 = wb.create_sheet('Ranking')
    ws2.append(['Cod', 'Apelido', 'Fat. Atual', 'Fat. Anterior', 'Delta %', 'Pedidos'])
    for v in ranking:
        ws2.append([v['codvend'], v['apelido'], v['fat_atual'], v['fat_anterior'], v['delta_percentual'] / 100, v.get('pedidos_atual', 0)])
    _apply_excel_header(ws2, 6)
    for cell in ws2['C'][1:]:
        cell.number_format = 'R$ #,##0.00'
    for cell in ws2['D'][1:]:
        cell.number_format = 'R$ #,##0.00'
    for cell in ws2['E'][1:]:
        cell.number_format = '0.0%'
    _auto_width(ws2)

    ws3 = wb.create_sheet('Top Produtos')
    ws3.append(['Cod', 'Descricao', 'Marca', 'Qtd Vendida', 'Faturamento', 'Estoque'])
    for p in produtos:
        ws3.append([p['codprod'], p['descrprod'], p.get('marca'), float(p['qtd_vendida']), float(p['faturamento']), int(p.get('estoque') or 0)])
    _apply_excel_header(ws3, 6)
    for cell in ws3['E'][1:]:
        cell.number_format = 'R$ #,##0.00'
    _auto_width(ws3)

    ws4 = wb.create_sheet('Inativos')
    ws4.append(['Cod', 'Cliente', 'Cidade', 'UF', 'Valor Anterior', 'Vendedor'])
    for c in inativos:
        ws4.append([c['codparc'], c['nomeparc'], c.get('cidade'), c.get('uf'), float(c['valor_anterior']), c.get('apelido')])
    _apply_excel_header(ws4, 6)
    for cell in ws4['E'][1:]:
        cell.number_format = 'R$ #,##0.00'
    _auto_width(ws4)

    ws5 = wb.create_sheet('Diario')
    ws5.append(['Data', 'Faturamento'])
    for r in diario:
        ws5.append([r['data'], float(r['faturamento'])])
    _apply_excel_header(ws5, 2)
    for cell in ws5['B'][1:]:
        cell.number_format = 'R$ #,##0.00'
    _auto_width(ws5)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _completo_pdf(db: AsyncSession) -> bytes:
    kpis = await get_kpis_repository(db)
    ranking = await get_vendedor_ranking_repository(db)
    produtos = await get_top_produtos_repository(db, limit=15)
    inativos = await get_clientes_inativos_repository(db)
    styles = getSampleStyleSheet()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30, bottomMargin=30)
    story: list = [
        Paragraph('<b>Relatorio Comercial Completo</b>', styles['Title']),
        Paragraph(datetime.now(ZoneInfo('America/Recife')).strftime('Gerado em %d/%m/%Y %H:%M'), styles['Italic']),
        Spacer(1, 12),
        Paragraph('<b>KPIs do mes</b>', styles['Heading2']),
    ]
    kpi_data = [
        ['Metrica', 'Valor'],
        ['Faturamento Atual', _fmt_brl(kpis['faturamento_atual'])],
        ['Faturamento Anterior', _fmt_brl(kpis['faturamento_anterior'])],
        ['Delta', _fmt_pct(kpis['delta_percentual'])],
        ['Pedidos Atual', str(kpis.get('pedidos_atual', 0))],
        ['Clientes Ativos', str(kpis.get('clientes_ativos', 0))],
    ]
    story.append(Table(kpi_data, repeatRows=1, style=_table_style()))
    story.append(Spacer(1, 16))

    story.append(Paragraph('<b>Ranking de Vendedores</b>', styles['Heading2']))
    rnk = [['Cod', 'Apelido', 'Fat. Atual', 'Delta %']]
    for v in ranking:
        rnk.append([v['codvend'], v['apelido'], _fmt_brl(v['fat_atual']), _fmt_pct(v['delta_percentual'])])
    story.append(Table(rnk, repeatRows=1, style=_table_style()))
    story.append(Spacer(1, 16))

    story.append(Paragraph('<b>Top 15 Produtos</b>', styles['Heading2']))
    prd = [['Cod', 'Descricao', 'Faturamento']]
    for p in produtos:
        prd.append([p['codprod'], (p['descrprod'] or '')[:45], _fmt_brl(p['faturamento'])])
    story.append(Table(prd, repeatRows=1, style=_table_style()))
    story.append(Spacer(1, 16))

    if inativos:
        story.append(Paragraph(f'<b>Clientes Inativos ({len(inativos)})</b>', styles['Heading2']))
        ina = [['Cod', 'Cliente', 'Valor Anterior']]
        for c in inativos[:30]:
            ina.append([c['codparc'], (c['nomeparc'] or '')[:40], _fmt_brl(c['valor_anterior'])])
        story.append(Table(ina, repeatRows=1, style=_table_style()))

    doc.build(story)
    return buf.getvalue()


_GENERATORS = {
    ('ranking_vendedores', 'pdf'): (_ranking_pdf, 'ranking_vendedores', PDF_MIME, 'pdf'),
    ('ranking_vendedores', 'excel'): (_ranking_xlsx, 'ranking_vendedores', XLSX_MIME, 'xlsx'),
    ('top_produtos', 'pdf'): (_produtos_pdf, 'top_produtos', PDF_MIME, 'pdf'),
    ('top_produtos', 'excel'): (_produtos_xlsx, 'top_produtos', XLSX_MIME, 'xlsx'),
    ('clientes_inativos', 'pdf'): (_inativos_pdf, 'clientes_inativos', PDF_MIME, 'pdf'),
    ('clientes_inativos', 'excel'): (_inativos_xlsx, 'clientes_inativos', XLSX_MIME, 'xlsx'),
    ('comparativo_diario', 'pdf'): (_diario_pdf, 'comparativo_diario', PDF_MIME, 'pdf'),
    ('comparativo_diario', 'excel'): (_diario_xlsx, 'comparativo_diario', XLSX_MIME, 'xlsx'),
    ('completo', 'pdf'): (_completo_pdf, 'relatorio_completo', PDF_MIME, 'pdf'),
    ('completo', 'excel'): (_completo_xlsx, 'relatorio_completo', XLSX_MIME, 'xlsx'),
}

VALID_TIPOS = sorted({t for (t, _) in _GENERATORS.keys()})
VALID_FORMATOS = ['pdf', 'excel']


async def gerar_relatorio(db: AsyncSession, tipo: str, formato: str) -> dict:
    '''Gera um relatorio e retorna dict com filename, mime_type e data_base64.'''
    key = (tipo, formato)
    if key not in _GENERATORS:
        raise ValueError(f'combinacao invalida: tipo={tipo} formato={formato}')
    generator, base, mime, ext = _GENERATORS[key]
    data = await generator(db)
    return {
        'filename': f'{base}_{_today_tag()}.{ext}',
        'mime_type': mime,
        'data_base64': base64.b64encode(data).decode('ascii'),
        'size_bytes': len(data),
    }
